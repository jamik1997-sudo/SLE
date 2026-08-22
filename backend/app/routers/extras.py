import time
import os
import json
import subprocess
from datetime import datetime
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, or_, func, text
from sqlalchemy.orm import Session, selectinload
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from app.database import DB_MAX_OVERFLOW, DB_POOL_SIZE, get_db
from app.models import (ActivityLog, Answer, Audit, AuditStatus, Employee, QuestionSetting, Region, Role, ScoreSetting, User, Visit, VisitTiming)
from app.security import current_user, require_roles
from app.timezone_utils import to_tashkent_naive
from pydantic import BaseModel



def _legacy_visit_text(value):
    """Normalize current/legacy visit text values for reports."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()

def _legacy_visit_goal_and_comment(visit=None, audit=None):
    """
    Backward-compatible lookup for visit goal/comment.
    Current fields are preferred; then known legacy aliases / JSON containers.
    """
    goal = ""
    comment = ""

    def pick(obj, names):
        if obj is None:
            return ""
        for name in names:
            try:
                if isinstance(obj, dict):
                    v = obj.get(name)
                else:
                    v = getattr(obj, name, None)
            except Exception:
                v = None
            v = _legacy_visit_text(v)
            if v:
                return v
        return ""

    goal_names = (
        "goal", "visit_goal", "purpose", "visit_purpose",
        "goal_text", "purpose_text", "target", "visit_target",
    )
    comment_names = (
        "comment", "visit_comment", "goal_comment",
        "purpose_comment", "comment_text", "visit_notes", "notes",
    )

    goal = pick(visit, goal_names)
    comment = pick(visit, comment_names)

    # Search JSON-ish legacy containers, if present.
    containers = []
    for obj in (visit, audit):
        if obj is None:
            continue
        for name in ("data", "payload", "draft", "draft_data", "meta", "metadata", "visit_data", "extra"):
            try:
                value = obj.get(name) if isinstance(obj, dict) else getattr(obj, name, None)
            except Exception:
                value = None
            if isinstance(value, str):
                try:
                    value = json.loads(value)
                except Exception:
                    value = None
            if isinstance(value, dict):
                containers.append(value)

    for data in containers:
        if not goal:
            goal = pick(data, goal_names)
        if not comment:
            comment = pick(data, comment_names)
        # Common nested visit object
        nested = data.get("visit") if isinstance(data, dict) else None
        if isinstance(nested, dict):
            if not goal:
                goal = pick(nested, goal_names)
            if not comment:
                comment = pick(nested, comment_names)

    # Some old builds stored initial visit fields directly on Audit.
    if not goal:
        goal = pick(audit, goal_names)
    if not comment:
        comment = pick(audit, comment_names)

    return goal or "—", comment or "—"


SLE_PROCESS_STARTED_AT = time.time()

router = APIRouter(prefix="/extras", tags=["extras"])

def allowed_regions(user): return {x.region_id for x in user.regions}

def scope(stmt, user):
    stmt = stmt.where(Audit.status != AuditStatus.cancelled)
    if user.role == Role.leader:
        stmt = stmt.where(Audit.region_id.in_(allowed_regions(user)))
    return stmt

def log(db, user, action, entity_type=None, entity_id=None, details=None):
    db.add(ActivityLog(user_id=user.id if user else None, action=action, entity_type=entity_type, entity_id=entity_id, details=details))

@router.get("/search")
def global_search(q: str = Query(min_length=1), db: Session = Depends(get_db), user: User = Depends(current_user)):
    term=f"%{q.strip()}%"
    estmt=select(Employee).options(selectinload(Employee.region)).where(Employee.is_active==True, or_(Employee.full_name.ilike(term), Employee.position.ilike(term)))
    if user.role==Role.leader: estmt=estmt.where(Employee.region_id.in_(allowed_regions(user)))
    employees=db.scalars(estmt.limit(20)).all()
    astmt=scope(select(Audit).options(selectinload(Audit.employee),selectinload(Audit.region),selectinload(Audit.auditor)).join(Audit.employee).join(Audit.region).where(or_(Employee.full_name.ilike(term),Region.name.ilike(term))),user)
    audits=db.scalars(astmt.limit(20)).all()
    return {"employees":[{"id":e.id,"name":e.full_name,"position":e.position,"region":e.region.name} for e in employees],"audits":[{"id":a.id,"date":a.audit_date,"employee":a.employee.full_name,"region":a.region.name,"percent":a.total_percent,"status":a.status.value} for a in audits]}

@router.get("/employees/{employee_id}")
def employee_card(employee_id: str, db: Session=Depends(get_db), user: User=Depends(current_user)):
    e=db.scalar(select(Employee).options(selectinload(Employee.region)).where(Employee.id==employee_id))
    if not e: raise HTTPException(404,"Сотрудник не найден")
    if user.role==Role.leader and e.region_id not in allowed_regions(user): raise HTTPException(403,"Нет доступа")
    rows=db.scalars(select(Audit).where(Audit.employee_id==employee_id,Audit.status==AuditStatus.completed).order_by(Audit.audit_date.desc())).all()
    avg=round(sum((a.total_percent or 0) for a in rows)/len(rows),1) if rows else 0
    trend=[{"date":str(a.audit_date),"percent":a.total_percent or 0,"level":a.level} for a in reversed(rows[:12])]
    return {"id":e.id,"full_name":e.full_name,"position":e.position,"region":e.region.name,"audits":len(rows),"average":avg,"last_result":rows[0].total_percent if rows else None,"trend":trend}

@router.get("/logs")
def logs(limit:int=100,db:Session=Depends(get_db),user:User=Depends(require_roles(Role.admin,Role.manager))):
    rows=db.scalars(select(ActivityLog).options(selectinload(ActivityLog.user)).order_by(ActivityLog.created_at.desc()).limit(min(limit,500))).all()
    return [{"id":x.id,"user":x.user.full_name if x.user else "Система","action":x.action,"entity_type":x.entity_type,"details":x.details,"created_at":x.created_at} for x in rows]

@router.get("/question-settings")
def questions(db:Session=Depends(get_db),user:User=Depends(current_user)):
    return db.scalars(select(QuestionSetting).order_by(QuestionSetting.sort_order)).all()

@router.put("/question-settings/{key}")
def update_question(key:str,payload:dict,db:Session=Depends(get_db),user:User=Depends(require_roles(Role.admin))):
    q=db.get(QuestionSetting,key)
    if not q: raise HTTPException(404,"Вопрос не найден")
    for f in ("text_ru","text_uz","section","step","weight","sort_order","is_active"):
        if f in payload: setattr(q,f,payload[f])
    log(db,user,"Изменил вопрос","question",key,q.text_ru)
    db.commit(); return {"saved":True}

@router.get("/score-settings")
def get_score_settings(db:Session=Depends(get_db),user:User=Depends(current_user)):
    s=db.get(ScoreSetting,1) or ScoreSetting(id=1)
    return {"confident_min":s.confident_min,"master_min":s.master_min}

@router.put("/score-settings")
def put_score_settings(payload:dict,db:Session=Depends(get_db),user:User=Depends(require_roles(Role.admin))):
    s=db.get(ScoreSetting,1)
    if not s: s=ScoreSetting(id=1); db.add(s)
    c=float(payload.get("confident_min",65)); m=float(payload.get("master_min",85))
    if not 0<c<m<=100: raise HTTPException(422,"Проверьте границы уровней")
    s.confident_min=c;s.master_min=m;log(db,user,"Изменил уровни оценки","settings","score",f"{c}/{m}");db.commit();return {"saved":True}

@router.post("/audit/{audit_id}/visit/{visit_number}/start")
def start_visit(audit_id:str,visit_number:int,db:Session=Depends(get_db),user:User=Depends(current_user)):
    a=db.get(Audit,audit_id)
    if not a: raise HTTPException(404,"Аудит не найден")
    if user.role==Role.leader and a.auditor_id!=user.id: raise HTTPException(403,"Нет доступа")
    t=db.scalar(select(VisitTiming).where(VisitTiming.audit_id==audit_id,VisitTiming.visit_number==visit_number))
    if not t: t=VisitTiming(audit_id=audit_id,visit_number=visit_number);db.add(t)
    if not t.started_at:t.started_at=datetime.utcnow()
    db.commit();return {"started_at":t.started_at}

@router.post("/audit/{audit_id}/visit/{visit_number}/end")
def end_visit(audit_id:str,visit_number:int,db:Session=Depends(get_db),user:User=Depends(current_user)):
    t=db.scalar(select(VisitTiming).where(VisitTiming.audit_id==audit_id,VisitTiming.visit_number==visit_number))
    if not t: t=VisitTiming(audit_id=audit_id,visit_number=visit_number,started_at=datetime.utcnow());db.add(t)
    t.ended_at=datetime.utcnow();db.commit();return {"ended_at":t.ended_at}

class OfflineTimingIn(BaseModel):
    started_at: datetime | None = None
    ended_at: datetime | None = None


@router.put("/audit/{audit_id}/visit/{visit_number}/offline-timing")
def offline_timing(audit_id:str,visit_number:int,payload:OfflineTimingIn,db:Session=Depends(get_db),user:User=Depends(current_user)):
    a=db.get(Audit,audit_id)
    if not a: raise HTTPException(404,"Аудит не найден")
    if user.role in (Role.leader, Role.auditor) and a.auditor_id!=user.id: raise HTTPException(403,"Нет доступа")
    t=db.scalar(select(VisitTiming).where(VisitTiming.audit_id==audit_id,VisitTiming.visit_number==visit_number))
    if not t: t=VisitTiming(audit_id=audit_id,visit_number=visit_number);db.add(t)
    if payload.started_at is not None and t.started_at is None: t.started_at=payload.started_at
    if payload.ended_at is not None: t.ended_at=payload.ended_at
    db.commit();return {"saved":True}


@router.get("/audit/{audit_id}/timings")
def timings(audit_id:str,db:Session=Depends(get_db),user:User=Depends(current_user)):
    rows=db.scalars(select(VisitTiming).where(VisitTiming.audit_id==audit_id).order_by(VisitTiming.visit_number)).all()
    return [{"visit_number":x.visit_number,"started_at":x.started_at,"ended_at":x.ended_at,"minutes":round((x.ended_at-x.started_at).total_seconds()/60,1) if x.started_at and x.ended_at else None} for x in rows]



@router.get("/system-status")
def system_status(db: Session = Depends(get_db), user: User = Depends(require_roles(Role.admin))):
    started = time.perf_counter()
    db_ok = True
    db_ms = None
    db_error = None
    try:
        db_started = time.perf_counter()
        db.execute(text("SELECT 1"))
        db_ms = round((time.perf_counter() - db_started) * 1000, 1)
    except Exception as exc:
        db_ok = False
        db_error = str(exc)[:240]

    try:
        l1, l5, l15 = os.getloadavg()
        load = {"1m": round(l1,2), "5m": round(l5,2), "15m": round(l15,2)}
    except Exception:
        load = None

    pool = None
    try:
        eng = db.get_bind()
        p = getattr(eng, "pool", None)
        if p is not None:
            size = p.size() if hasattr(p, "size") else DB_POOL_SIZE
            checked_out = p.checkedout() if hasattr(p, "checkedout") else 0
            max_overflow = DB_MAX_OVERFLOW
            capacity = max(1, size + max_overflow)
            pool = {
                "size": size,
                "max_overflow": max_overflow,
                "capacity": capacity,
                "checked_out": checked_out,
                "overflow": p.overflow() if hasattr(p,"overflow") else None,
                "checked_in": p.checkedin() if hasattr(p,"checkedin") else None,
                "utilization_percent": round(checked_out / capacity * 100, 1),
                "high_load": checked_out >= capacity,
            }
    except Exception:
        pass

    return {
        "status": "ok" if db_ok else "degraded",
        "api": {"ok": True, "response_ms": round((time.perf_counter()-started)*1000,1), "pid": os.getpid(), "uptime_seconds": int(time.time()-SLE_PROCESS_STARTED_AT)},
        "database": {"ok": db_ok, "response_ms": db_ms, "error": db_error},
        "load": load,
        "pool": pool,
    }

@router.post("/system-update", status_code=202)
def system_update(user: User = Depends(require_roles(Role.admin))):
    """Start the fixed, root-managed updater; no user command is accepted."""
    check = subprocess.run(
        ["/usr/bin/systemctl", "is-active", "sle-update.service"],
        capture_output=True,
        text=True,
        timeout=5,
        check=False,
    )
    if check.stdout.strip() in {"active", "activating"}:
        return {"status": "already_running", "message": "Обновление уже выполняется"}

    try:
        started = subprocess.run(
            ["sudo", "-n", "/usr/bin/systemctl", "start", "sle-update.service"],
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise HTTPException(status_code=503, detail=f"Не удалось запустить обновление: {exc}") from exc

    if started.returncode != 0:
        error = (started.stderr or started.stdout or "systemctl error").strip()[:300]
        raise HTTPException(status_code=503, detail=f"Не удалось запустить обновление: {error}")

    return {"status": "started", "message": "Обновление backend запущено"}

@router.get("/questionnaire-report")
def questionnaire_report(
    include_details: bool = False,
    limit: int = 1000,
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    stmt = (
        select(Audit)
        .options(
            selectinload(Audit.employee),
            selectinload(Audit.region),
            selectinload(Audit.auditor),
            selectinload(Audit.answers),
            selectinload(Audit.visits),
        )
        .order_by(Audit.audit_date.desc(), Audit.last_saved_at.desc())
    )
    stmt = scope(stmt, user)
    audits = db.scalars(stmt.limit(min(max(limit, 1), 3000))).all()
    qrows = db.scalars(select(QuestionSetting).where(QuestionSetting.is_active == True).order_by(QuestionSetting.sort_order)).all()
    questions = [{
        "key": q.key, "section": q.section, "step": q.step, "weight": q.weight, "text": q.text_ru
    } for q in qrows]
    stats = {q["key"]: {**q, "filled": 0, "ones": 0, "zeros": 0} for q in questions}
    status_counts = {"draft": 0, "in_progress": 0, "completed": 0, }
    details = []
    total_answers = 0
    for audit in audits:
        status_counts[audit.status.value] = status_counts.get(audit.status.value, 0) + 1
        for answer in audit.answers:
            item = stats.get(answer.question_key)
            if not item:
                continue
            item["filled"] += 1
            total_answers += 1
            if answer.answer_value == "1": item["ones"] += 1
            elif answer.answer_value == "0": item["zeros"] += 1
            if include_details and len(details) < 30000:
                details.append({
                    "audit_id": audit.id,
                    "audit_date": str(audit.audit_date),
                    "status": audit.status.value,
                    "region": audit.region.name,
                    "employee": audit.employee.full_name,
                    "auditor": audit.auditor.full_name,
                    "visit_number": answer.visit_number,
                    "section": item["section"],
                    "question": item["text"],
                    "answer": answer.answer_value,
                    "updated_at": answer.updated_at.isoformat() if answer.updated_at else None,
                })
    summary = []
    audit_count = len(audits)
    for q in questions:
        item = stats[q["key"]]
        expected_per_audit = 1 if q["step"] in (0, 8) else 5
        expected = audit_count * expected_per_audit
        item["expected"] = expected
        item["completion_percent"] = round(item["filled"] / expected * 100, 1) if expected else 0
        item["success_percent"] = round(item["ones"] / item["filled"] * 100, 1) if item["filled"] else 0
        summary.append(item)
    return {
        "audit_count": audit_count,
        "total_answers": total_answers,
        "status_counts": status_counts,
        "questions": summary,
        "details": details if include_details else [],
    }


def _xlsx_response(workbook: Workbook, filename: str):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="FFD600")
    header_font = Font(bold=True, color="111111")
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 500) + 1):
            value = ws.cell(row, col).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 42)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


@router.get("/export/audit-report.xlsx")
def export_audits_xlsx(
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    stmt = (
        select(Audit)
        .options(selectinload(Audit.employee), selectinload(Audit.region), selectinload(Audit.auditor), selectinload(Audit.visits))
        .order_by(Audit.audit_date.desc(), Audit.last_saved_at.desc())
    )
    audits = db.scalars(scope(stmt, user)).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Аудиты"
    ws.append(["Дата", "Сотрудник", "Регион", "Оценивающий", "Статус", "Результат, %", "Уровень", "Торговые точки", "Цель визита", "Комментарий", "Координаты ТТ", "Начат", "Отправлен"])
    status_names = {"draft":"Черновик", "in_progress":"В процессе", "completed":"Завершён"}
    for a in audits:
        shop_names = "; ".join(f"{v.visit_number}. {v.shop_code or '—'}" for v in a.visits)
        coordinates = "; ".join(f"{v.visit_number}. {v.latitude},{v.longitude}" for v in a.visits if v.latitude is not None and v.longitude is not None)
        goals = "; ".join(f"{v.visit_number}. {(v.goal or '').strip()}" for v in sorted(a.visits, key=lambda x: x.visit_number) if (v.goal or '').strip())
        comments = "; ".join(f"{v.visit_number}. {(v.comment or '').strip()}" for v in sorted(a.visits, key=lambda x: x.visit_number) if (v.comment or '').strip())
        ws.append([
            a.audit_date,
            a.employee.full_name if a.employee else "",
            a.region.name if a.region else "",
            a.auditor.full_name if a.auditor else "",
            status_names.get(a.status.value, a.status.value),
            a.total_percent,
            a.level or "",
            shop_names,
            goals,
            comments,
            coordinates,
            to_tashkent_naive(a.started_at),
            to_tashkent_naive(a.submitted_at),
        ])
    for cell in ws["A"][1:]: cell.number_format = "dd.mm.yyyy"
    for col in (12, 13):
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell: c.number_format = "dd.mm.yyyy hh:mm"
    _style_sheet(ws)
    loc = wb.create_sheet("Торговые точки")
    loc.append(["ID аудита", "Дата", "Регион", "Сотрудник", "Оценивающий", "Визит", "Код ТТ", "Цель визита", "Комментарий", "Широта", "Долгота", "Координаты", "Ссылка на карту"])
    for a in audits:
        for v in a.visits:
            coords = f"{v.latitude},{v.longitude}" if v.latitude is not None and v.longitude is not None else ""
            loc.append([a.id, a.audit_date, a.region.name if a.region else "", a.employee.full_name if a.employee else "", a.auditor.full_name if a.auditor else "", v.visit_number, v.shop_code, v.goal or "", v.comment or "", v.latitude, v.longitude, coords, f"https://maps.google.com/?q={coords}" if coords else ""])
            if coords:
                cell = loc.cell(loc.max_row, 13); cell.hyperlink = cell.value; cell.style = "Hyperlink"
    _style_sheet(loc)
    return _xlsx_response(wb, "audit-report.xlsx")


@router.get("/export/detailed-report.xlsx")
def export_questionnaire_xlsx(
    limit: int = 3000,
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    stmt = (
        select(Audit)
        .options(
            selectinload(Audit.employee), selectinload(Audit.region),
            selectinload(Audit.auditor), selectinload(Audit.answers), selectinload(Audit.visits),
        )
        .order_by(Audit.audit_date.desc(), Audit.last_saved_at.desc())
    )
    audits = db.scalars(scope(stmt, user).limit(min(max(limit, 1), 3000))).all()
    qrows = db.scalars(select(QuestionSetting).where(QuestionSetting.is_active == True).order_by(QuestionSetting.sort_order)).all()
    qmap = {q.key: q for q in qrows}
    stats = {q.key: {"q": q, "filled": 0, "ones": 0, "zeros": 0} for q in qrows}

    wb = Workbook()
    ws = wb.active
    ws.title = "Детальный отчет"
    ws.append(["ID аудита", "Дата", "Статус", "Регион", "Сотрудник", "Оценивающий", "Визит", "Код ТТ", "Цель визита", "Комментарий визита", "Широта", "Долгота", "Раздел", "Вопрос", "Ответ", "Комментарий к ответу", "Обновлено"])
    status_names = {"draft":"Черновик", "in_progress":"В процессе", "completed":"Завершён"}
    for audit in audits:
        for answer in audit.answers:
            q = qmap.get(answer.question_key)
            if not q: continue
            st = stats[q.key]
            st["filled"] += 1
            if answer.answer_value == "1": st["ones"] += 1
            elif answer.answer_value == "0": st["zeros"] += 1
            visit = next((v for v in audit.visits if v.visit_number == answer.visit_number), None)
            ws.append([
                audit.id, audit.audit_date, status_names.get(audit.status.value, audit.status.value),
                audit.region.name if audit.region else "", audit.employee.full_name if audit.employee else "",
                audit.auditor.full_name if audit.auditor else "", answer.visit_number,
                visit.shop_code if visit else "", _legacy_visit_goal_and_comment(visit, locals().get("audit"))[0] if visit else "", visit.comment if visit else "",
                visit.latitude if visit else None, visit.longitude if visit else None,
                q.section, q.text_ru, answer.answer_value, answer.comment or "", to_tashkent_naive(answer.updated_at),
            ])
    for cell in ws["B"][1:]: cell.number_format = "dd.mm.yyyy"
    for cell in ws["Q"][1:]: cell.number_format = "dd.mm.yyyy hh:mm"
    _style_sheet(ws)

    sm = wb.create_sheet("Сводка")
    sm.append(["Раздел", "Вопрос", "Вес", "Заполнено", "Ожидалось", "Заполнение, %", "Ответ 1", "Ответ 0", "Выполнение, %"])
    audit_count = len(audits)
    for q in qrows:
        st = stats[q.key]
        expected_per_audit = 1 if q.step in (0, 8) else 5
        expected = audit_count * expected_per_audit
        completion = round(st["filled"] / expected * 100, 1) if expected else 0
        success = round(st["ones"] / st["filled"] * 100, 1) if st["filled"] else 0
        sm.append([q.section, q.text_ru, q.weight, st["filled"], expected, completion, st["ones"], st["zeros"], success])
    for col in (6, 9):
        for cell in sm.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell: c.number_format = '0.0"%"'
    _style_sheet(sm)
    return _xlsx_response(wb, "detailed-audit-report.xlsx")
